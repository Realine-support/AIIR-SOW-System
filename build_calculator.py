import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── SHEET 1: Calculator ──────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Calculator'

bold    = Font(bold=True)
bold14  = Font(bold=True, size=14)
bold12  = Font(bold=True, size=12)
red_fill    = PatternFill('solid', fgColor='FFE0E0')
blue_fill   = PatternFill('solid', fgColor='E0F0FF')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
header_fill = PatternFill('solid', fgColor='1a73e8')
white_font  = Font(bold=True, color='FFFFFF')
red_italic  = Font(color='CC0000', italic=True)

def s(row, col, value, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    return cell

# Title
s(1,1,'AIIR Pricing Calculator', bold14)
s(2,1,'Selected Tier');  s(2,2,'IGNITE')
s(3,1,'Engagement Date'); s(3,2,'')

# Pricing Inputs section
s(13,1,'-- PRICING INPUTS --', bold)
s(14,1,'Tier');                 s(14,2,'IGNITE')
s(15,1,'Coaching Bill Rate');   s(15,2,350,  fill=blue_fill)
s(16,1,'Margin');               s(16,2,0.65, fill=red_fill)
s(16,3,'<-- ALWAYS 0.65 -- do not change', red_italic)
s(17,1,'CZ Rate IGNITE');       s(17,2,75)
s(18,1,'CZ Rate ASCENT/VISTA'); s(18,2,50)

# ROADMAP reference (rows 20-33)
s(20,1,'-- ROADMAP TIER (reference) --', bold)
s(21,1,'Session', bold); s(21,2,'Hours', bold)
roadmap = [
    ('Initial Coaching Session', 0.5),
    ('Stakeholder 1: ASM', 1.0),
    ('Dev History Interview', 2.0),
    ('Interview-Based 360', 3.75),
    ('Assessment Feedback', 2.0),
    ('Dev Planning Session', 1.0),
    ('Stakeholder 2: SDP', 1.0),
    ('Implementation Sessions', 5.0),
    ('Stakeholder 3: Midpoint', 1.0),
    ('Stakeholder 4: Wrap-Up', 1.0),
]
for i,(name,hrs) in enumerate(roadmap):
    s(22+i,1,name); s(22+i,2,hrs)
s(32,1,'Total Hours (ROADMAP)', bold)
ws['B32'] = '=SUM(B22:B31)'
ws['B32'].font = bold
s(25,4,'CZ Fee (months)'); s(25,5,6)

# IGNITE active calculation area (rows 35-50)
s(35,1,'-- IGNITE TIER (ACTIVE CALCULATION AREA) --', bold)
s(36,1,'Session', bold); s(36,2,'Hours', bold); s(36,5,'Fee/Months', bold)

ignite = [
    ('Initial Coaching Session',  0.5,  'fixed'),
    ('Stakeholder 1: ASM',        1.0,  'Default: 1hr | Min: 30min'),
    ('Dev History Interview',     2.0,  'Default: 2hr | Min: 90min'),
    ('Interview-Based 360',       6.0,  'Default: 8x45min=6hr | Min: 5x45min=3.75hr'),
    ('Assessment Feedback',       2.0,  'Default: 2hr | Min: 90min'),
    ('Dev Planning Session',      1.0,  'Default: 1hr | Min: 0hr'),
    ('Stakeholder 2: SDP',        1.0,  'Default: 1hr | Min: 30min'),
    ('Implementation Sessions',   8.0,  'Default: 8x1hr'),
    ('Stakeholder 3: Midpoint',   1.0,  'Default: 1hr | Min: 30min'),
    ('Stakeholder 4: Wrap-Up',    1.0,  'Default: 1hr | Min: 30min'),
]
for i,(name,hrs,note) in enumerate(ignite):
    s(37+i,1,name); s(37+i,2,hrs); s(37+i,3,note)

# CZ fee inputs
s(37,4,'CZ Fee (months)');   s(37,5,9)   # E37 -- n8n writes 7-9
s(38,4,'CZ Rate/mo');        s(38,6,75)  # F38 -- n8n writes tier rate ($75 or $50)

# Coaching cost calc (B47:B50)
s(47,1,'Total Coaching Hours', bold)
ws['B47'] = '=SUM(B37:B46)'
ws['B47'].font = bold

s(48,1,'Total Coach Cost')
ws['B48'] = '=B47*B15'

s(49,1,'PM Fee (12%)')
ws['B49'] = '=B48*0.12'

s(50,1,'Total Services (no margin)')
ws['B50'] = '=B48+B49'

# OUTPUT section (columns G/H)
s(44,7,'CZ Fee')
ws['H44'] = '=E37*F38'

s(45,7,'Fixed Assessment Fees', None, blue_fill)
s(45,8,450, None, blue_fill)   # H45 -- n8n writes tier-specific value (default 450 = IGNITE)
ws.cell(45,9).value = '<-- n8n writes this per tier'
ws.cell(45,9).font = red_italic

s(47,7,'Services incl. Margin')
ws['H47'] = '=B50/(1-B16)'

s(48,7,'Total per Participant', bold)
ws['H48'] = '=H44+H45+H47'
ws['H48'].font = bold

s(49,7,'Number of Participants', bold, blue_fill)
ws['H49'] = ''   # n8n writes here FIRST

s(50,7,'TOTAL ENGAGEMENT PRICE', bold)
ws['H50'] = '=IF(H49="",H48,H48*H49)'
ws['H50'].font = bold12
ws['H50'].fill = yellow_fill

# ASCENT reference (rows 53-64)
s(53,1,'-- ASCENT TIER (reference) --', bold)
ascent = [
    ('Initial Coaching Session',  0.5),
    ('Stakeholder 1: ASM',        1.0),
    ('Dev History Interview',     2.0),
    ('Interview-Based 360',       9.0),
    ('Assessment Feedback',       2.0),
    ('Dev Planning Session',      1.0),
    ('Stakeholder 2: SDP',        1.0),
    ('Implementation Sessions',  12.0),
    ('Stakeholder 3: Midpoint',   1.0),
    ('Stakeholder 4: Wrap-Up',    1.0),
]
for i,(name,hrs) in enumerate(ascent):
    s(54+i,1,name); s(54+i,2,hrs)
s(58,4,'CZ Fee (months)'); s(58,5,12)

# SPARK I reference (rows 70-81)
s(70,1,'-- SPARK I TIER (reference) --', bold)
spark1 = [
    ('Initial Coaching Session',               0.5),
    ('Stakeholder 1: ASM',                     1.0),
    ('Dev History (Survey $60 - NO interview)',   0),
    ('Interview-Based 360',                   3.75),
    ('Assessment Feedback',                    2.0),
    ('Dev Planning Session',                   1.0),
    ('Stakeholder 2: SDP',                     1.0),
    ('Implementation Sessions',                4.0),
    ('Stakeholder 3: Midpoint',                1.0),
    ('Stakeholder 4: Wrap-Up',                 1.0),
]
for i,(name,hrs) in enumerate(spark1):
    s(71+i,1,name); s(71+i,2,hrs)
s(76,4,'CZ Fee (months)'); s(76,5,4)

# SPARK II reference (rows 88-99)
s(88,1,'-- SPARK II TIER (reference) --', bold)
spark2 = [
    ('Initial Coaching Session',  0.5),
    ('Stakeholder 1: ASM',        1.0),
    ('Dev History Interview',     2.0),
    ('Interview-Based 360',      3.75),
    ('Assessment Feedback',       2.0),
    ('Dev Planning Session',      1.0),
    ('Stakeholder 2: SDP',        1.0),
    ('Implementation Sessions',   5.0),
    ('Stakeholder 3: Midpoint',   1.0),
    ('Stakeholder 4: Wrap-Up',    1.0),
]
for i,(name,hrs) in enumerate(spark2):
    s(89+i,1,name); s(89+i,2,hrs)
s(94,4,'CZ Fee (months)'); s(94,5,5)

# AIIR VISTA reference (rows 105-116)
s(105,1,'-- AIIR VISTA TIER (reference) --', bold)
vista = [
    ('Initial Coaching Session',  0.5),
    ('Stakeholder 1: ASM',        1.0),
    ('Dev History Interview',     2.0),
    ('Interview-Based 360',       6.0),
    ('Assessment Feedback',       2.0),
    ('Dev Planning Session',      1.0),
    ('Stakeholder 2: SDP',        1.0),
    ('Implementation Sessions',  10.0),
    ('Stakeholder 3: Midpoint',   1.0),
    ('Stakeholder 4: Wrap-Up',    1.0),
]
for i,(name,hrs) in enumerate(vista):
    s(106+i,1,name); s(106+i,2,hrs)
s(111,4,'CZ Fee (months)'); s(111,5,12)

# Column widths
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 36
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 22
ws.freeze_panes = 'A4'

# ─── SHEET 2: Tracker ─────────────────────────────────────────────────────────
ts = wb.create_sheet('Tracker')
headers = [
    'engagement_id','created_at','transcript_filename','client_company',
    'coachee_name','hubspot_deal_id','coach_name','tier','bill_rate',
    'num_participants','price_per_participant','total_engagement_price',
    'payment_structure','net_days','pricing_approval_status',
    'pricing_approved_by','pricing_approved_at','rationale_doc_url',
    'calculator_url','sow_doc_url','sow_approval_status',
    'sow_approved_by','sow_approved_at','client_email',
    'sent_at','archive_folder_url','full_variables_json'
]
for col, h in enumerate(headers, 1):
    cell = ts.cell(row=1, column=col, value=h)
    cell.font = white_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    ts.column_dimensions[get_column_letter(col)].width = 22
ts.column_dimensions[get_column_letter(len(headers))].width = 50
ts.freeze_panes = 'A2'

wb.save('D:/AIIR/AIIR_Pricing_Calculator.xlsx')
print('Saved: D:/AIIR/AIIR_Pricing_Calculator.xlsx')
